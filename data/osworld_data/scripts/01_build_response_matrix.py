#!/usr/bin/env python3
"""
Build OSWorld response matrices from HuggingFace verified trajectory data.

OSWorld (NeurIPS 2024) evaluates AI agents on 369 real-world computer tasks across
Ubuntu with real applications (Chrome, LibreOffice, VS Code, Thunderbird, etc.).

Data sources:
  - os-world.github.io/static/data/osworld_verified_results.xlsx
      Verified leaderboard with domain-level aggregate scores for ~55 models
  - os-world.github.io/static/data/self_reported_results.xlsx
      Self-reported results (screenshot, a11y tree, etc. modalities)
  - huggingface.co/datasets/xlangai/ubuntu_osworld_verified_trajs
      Per-task result.txt files inside trajectory zip archives
      Each result.txt contains a float score (0.0 or 1.0 typically)
  - github.com/xlang-ai/OSWorld evaluation_examples/test_all.json
      Master task list with 369 UUIDs organized by domain

Strategy:
  Uses HTTP range requests (via remotezip) to extract only the tiny result.txt
  files from multi-GB trajectory zips, avoiding full downloads.

Outputs:
  - response_matrix.csv: Continuous (0-1) score matrix (models x tasks)
  - response_matrix_binary.csv: Binary pass/fail (1 if score >= 0.99, else 0)
  - model_summary.csv: Per-model summary statistics
  - task_metadata.csv: Task UUID to domain mapping
  - leaderboard_verified.csv: Parsed verified leaderboard from Excel
  - leaderboard_self_reported.csv: Parsed self-reported leaderboard from Excel
"""

import json
import os
import re
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR = BASE_DIR / "raw"
PROCESSED_DIR = BASE_DIR / "processed"

HF_BASE = (
    "https://huggingface.co/datasets/xlangai/"
    "ubuntu_osworld_verified_trajs/resolve/main/"
)
EXCEL_VERIFIED_URL = (
    "https://os-world.github.io/static/data/osworld_verified_results.xlsx"
)
EXCEL_SELF_REPORTED_URL = (
    "https://os-world.github.io/static/data/self_reported_results.xlsx"
)
TASK_LIST_URL = (
    "https://raw.githubusercontent.com/xlang-ai/OSWorld/"
    "main/evaluation_examples/test_all.json"
)

# Known domains in OSWorld
DOMAINS = [
    "chrome", "gimp", "libreoffice_calc", "libreoffice_impress",
    "libreoffice_writer", "multi_apps", "os", "thunderbird", "vlc", "vs_code",
]

# ---------------------------------------------------------------------------
# Zip-to-model mapping
# Each entry: (zip_filename, model_label, step_count)
# Some zips contain multiple runs (turns); we average them.
# ---------------------------------------------------------------------------
ZIP_MODEL_MAP = [
    # --- OpenAI ---
    ("o3_15steps.zip", "o3", 15),
    ("o3_50steps.zip", "o3", 50),
    ("o3_100steps.zip", "o3", 100),
    # --- Claude ---
    ("claude-3-7-sonnet-20250219-15steps.zip",
     "claude-3-7-sonnet-20250219", 15),
    ("claude-3-7-sonnet-20250219-50steps.zip",
     "claude-3-7-sonnet-20250219", 50),
    ("claude-3-7-sonnet-20250219-100steps.zip",
     "claude-3-7-sonnet-20250219", 100),
    ("claude-4-sonnet-20250514-15steps.zip",
     "claude-4-sonnet-20250514", 15),
    ("claude-4-sonnet-20250514-50steps.zip",
     "claude-4-sonnet-20250514", 50),
    ("claude-4-sonnet-20250514-100steps.zip",
     "claude-4-sonnet-20250514", 100),
    ("claude-sonnet-4-5-20250929_15steps.zip",
     "claude-sonnet-4-5-20250929", 15),
    ("claude-sonnet-4-5-20250929_50steps.zip",
     "claude-sonnet-4-5-20250929", 50),
    ("claude-sonnet-4-5-20250929_100steps.zip",
     "claude-sonnet-4-5-20250929", 100),
    # --- ByteDance ---
    ("doubao-1-5-thinking-vision-pro-250428-15step.zip",
     "doubao-1-5-thinking-vision-pro-250428", 15),
    ("doubao-1-5-thinking-vision-pro-250428-100step.zip",
     "doubao-1-5-thinking-vision-pro-250428", 100),
    ("UI-TARS-0717-15step.zip", "UI-TARS-250705", 15),
    ("UI-TARS-0717-100step.zip", "UI-TARS-250705", 100),
    ("uitars-72b-dpo-15step.zip", "uitars-72b-dpo", 15),
    ("uitars-72b-dpo-50step.zip", "uitars-72b-dpo", 50),
    ("uitars-72b-dpo-100step.zip", "uitars-72b-dpo", 100),
    ("results_UI-TARS-2-2509_100steps.zip", "UI-TARS-2-2509", 100),
    # --- UI-TARS-1.5-7b (multiple runs in separate zips) ---
    ("uitars15-7b-15step.zip", "uitars-1.5-7b_run1", 15),
    ("uitars15-7b-15step-1.zip", "uitars-1.5-7b_run2", 15),
    ("uitars15-7b-50step.zip", "uitars-1.5-7b_run1", 50),
    ("uitars15-7b-50step-1.zip", "uitars-1.5-7b_run2", 50),
    ("uitars15-7b-100step.zip", "uitars-1.5-7b_run1", 100),
    ("uitars15-7b-100step-1.zip", "uitars-1.5-7b_run2", 100),
    # --- OpenCUA (multi-turn zips: each zip has turn_1/turn_2/turn_3) ---
    ("opencua_agent-opencua_7b-cot_l2-action_history-3image-Ubuntu-15steps.zip",
     "opencua-7b", 15),
    ("opencua_agent-opencua_7b-cot_l2-action_history-3image-Ubuntu-50steps.zip",
     "opencua-7b", 50),
    ("opencua_agent-opencua_7b-cot_l2-action_history-3image-Ubuntu-100steps.zip",
     "opencua-7b", 100),
    ("opencua_agent-opencua_32b-cot_l2-action_history-3image-Ubuntu-15steps.zip",
     "opencua-32b", 15),
    ("opencua_agent-opencua_32b-cot_l2-action_history-3image-Ubuntu-50steps.zip",
     "opencua-32b", 50),
    ("opencua_agent-opencua_32b-cot_l2-action_history-3image-Ubuntu-100steps.zip",
     "opencua-32b", 100),
    ("opencua_agent-opencua_a3b-cot_l2-action_history-3image-Ubuntu-15step.zip",
     "opencua-a3b", 15),
    ("opencua_agent-opencua_a3b-cot_l2-action_history-3image-Ubuntu-50step.zip",
     "opencua-a3b", 50),
    ("opencua_agent-opencua_a3b-cot_l2-action_history-3image-Ubuntu-100step.zip",
     "opencua-a3b", 100),
    ("opencua_agent-opencua_qwen2_7b-cot_l2-action_history-3image-Ubuntu-15step.zip",
     "opencua-qwen2-7b", 15),
    ("opencua_agent-opencua_qwen2_7b-cot_l2-action_history-3image-Ubuntu-50step.zip",
     "opencua-qwen2-7b", 50),
    ("opencua_agent-opencua_qwen2_7b-cot_l2-action_history-3image-Ubuntu-100step.zip",
     "opencua-qwen2-7b", 100),
    # --- Qwen ---
    ("qwen2.5-vl-32b-instruct_15step.zip",
     "qwen2.5-vl-32b-instruct", 15),
    ("qwen2.5-vl-32b-instruct_100step.zip",
     "qwen2.5-vl-32b-instruct", 100),
    ("qwen2.5-vl-72b-instruct_15step.zip",
     "qwen2.5-vl-72b-instruct", 15),
    ("qwen2.5-vl-72b-instruct_100step.zip",
     "qwen2.5-vl-72b-instruct", 100),
    # --- Kimi ---
    ("kimi-k25.zip", "Kimi K2.5", 100),
    ("kimi-vl-a3b-15step.zip", "kimi-vl-a3b", 15),
    ("kimi-vl-a3b-100step.zip", "kimi-vl-a3b", 100),
    # --- AutoGLM ---
    ("autoglm_15steps.zip", "autoglm-os-9b", 15),
    ("autoglm_50steps.zip", "autoglm-os-9b", 50),
    ("results_autoglm_v_15steps.zip", "autoglm-os-9b-20250925", 15),
    ("results_autoglm_v_50steps.zip", "autoglm-os-9b-20250925", 50),
    # --- Agent frameworks ---
    ("results_agent_s2_o3_15steps.zip", "agent s2.5 w/ o3", 15),
    ("results_agent_s2_gemini_15steps.zip",
     "agent s2 w/ gemini-2.5-pro", 15),
    ("results_gemini_50_steps_aws.zip",
     "agent s2 w/ gemini-2.5-pro", 50),
    ("o3_gta1_50steps.zip", "GTA1 w/ o3", 50),
    ("o3_gta1_100steps.zip", "GTA1 w/ o3", 100),
    # --- Jedi ---
    ("jedi-7b-4o-15steps.zip", "Jedi-7B w/ gpt-4o", 15),
    ("jedi-7b-4o-50steps.zip", "Jedi-7B w/ gpt-4o", 50),
    ("jedi-7b-4o-100steps.zip", "Jedi-7B w/ gpt-4o", 100),
    ("jedi-7b-o3-15steps.zip", "Jedi-7B w/ o3", 15),
    ("jedi-7b-o3-50steps.zip", "Jedi-7B w/ o3", 50),
    ("jedi-7b-o3-100steps.zip", "Jedi-7B w/ o3", 100),
    # --- EvoCUA ---
    ("evocua_20260105.zip", "EvoCUA-20260105", 50),
    ("evocua_8b_20260105.zip", "EvoCUA-8B-20260105", 50),
    # --- Maestro ---
    ("results_maestro_100steps_results_only.zip",
     "Agentic-Lybic-Maestro", 100),
    ("results_maestro_50steps_2_results_only.zip",
     "Agentic-Lybic-Maestro", 50),
    # --- Mano / DeepMiner ---
    ("results_mano_100steps.zip", "DeepMiner-Mano-72B", 100),
    # --- DART ---
    ("result_aws_dart_gui_20260122.zip", "DART-GUI-7B-0924", 30),
    # --- Others ---
    ("results_hippo_agent.zip", "HIPPO Agent w/Opus 4.5", 100),
    ("results_seed18.zip", "Seed-1.8", 100),
    ("results_gbox_15steps.zip", "GBOX Agent", 15),
    ("results_opus4p5_single_agent.zip",
     "UiPath Screen Agent w/ Opus 4.5", 100),
    ("uipath_gpt_5_50steps.zip",
     "UiPath Screen Agent w/ GPT-5", 50),
    ("mobile-agent-v3-gui-owl-7b.zip", "gui-owl-7b", 15),
    ("mobileagent_v3.zip",
     "mobile-agent-v3 w/ gui-owl-32b", 50),
    ("results_bbon_10_72_6.zip",
     "agent s3 w/ GPT-5 bBoN (N=10)", 100),
    ("results_tianxi_action_7b.zip", "TianXi-Action-7B", 50),
    ("results_omnitars_100_steps.zip",
     "qwen3-vl-flash-2025-10-25", 100),
    ("results_agi0_50steps.zip", "aworldGUIAgent-v1", 50),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

UUID_PATTERN = re.compile(
    r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}"
)


def download_file(url, dest):
    """Download a file if not already present."""
    if dest.exists() and dest.stat().st_size > 0:
        print(f"  Already exists: {dest.name}")
        return
    import urllib.request
    print(f"  Downloading {dest.name} ...")
    urllib.request.urlretrieve(url, str(dest))
    print(f"  Saved ({dest.stat().st_size / 1024:.1f} KB)")


def extract_results_from_zip(zip_url, max_retries=5):
    """Extract per-task results from a remote zip via HTTP range requests.

    Returns a list of dicts, one per run/turn found:
      [{"domain/uuid": score, ...}, ...]

    Each dict maps "domain/uuid" to a float score.
    """
    import remotezip

    for attempt in range(max_retries):
        try:
            rz = remotezip.RemoteZip(zip_url)
            names = rz.namelist()
            result_files = [
                n for n in names
                if n.endswith("result.txt") and "__MACOSX" not in n
            ]

            if not result_files:
                rz.close()
                return []

            # Group by run/turn prefix (everything before the domain)
            runs = {}
            for rf in result_files:
                # Find the domain and UUID in the path
                match = None
                for domain in DOMAINS:
                    idx = rf.find(f"/{domain}/")
                    if idx >= 0:
                        prefix = rf[:idx]
                        uuid_match = UUID_PATTERN.search(
                            rf[idx + len(domain) + 2:]
                        )
                        if uuid_match:
                            task_key = f"{domain}/{uuid_match.group()}"
                            match = (prefix, task_key)
                            break
                    # Handle case where domain is at start of path
                    if rf.startswith(f"{domain}/"):
                        uuid_match = UUID_PATTERN.search(
                            rf[len(domain) + 1:]
                        )
                        if uuid_match:
                            prefix = ""
                            task_key = f"{domain}/{uuid_match.group()}"
                            match = (prefix, task_key)
                            break

                if match:
                    prefix, task_key = match
                    if prefix not in runs:
                        runs[prefix] = {}
                    # Read the score
                    try:
                        content = rz.read(rf).decode("utf-8").strip()
                        score = float(content)
                        runs[prefix][task_key] = score
                    except (ValueError, UnicodeDecodeError):
                        pass

            rz.close()
            return list(runs.values())

        except Exception as exc:
            if attempt < max_retries - 1:
                # Exponential backoff: 10, 30, 60, 120 seconds
                wait = 10 * (2 ** attempt)
                print(
                    f"    Retry {attempt + 1}/{max_retries} "
                    f"in {wait}s after: {exc}"
                )
                time.sleep(wait)
            else:
                print(
                    f"    FAILED after {max_retries} attempts: "
                    f"{exc}"
                )
                return []


def load_task_list():
    """Load the master task list from GitHub or local cache.

    Returns dict: {domain: [uuid1, uuid2, ...]}
    """
    cache_path = RAW_DIR / "test_all.json"
    if not cache_path.exists():
        download_file(TASK_LIST_URL, cache_path)

    with open(cache_path, "r") as f:
        data = json.load(f)

    # Data format: {"domain": ["uuid1", "uuid2", ...]}
    return data


def parse_verified_excel(excel_path):
    """Parse the verified leaderboard Excel into a DataFrame."""
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb["Eval Results"]

    headers = [
        cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h is not None and i < len(row):
                row_dict[h] = row[i]
        rows.append(row_dict)

    return pd.DataFrame(rows)


def parse_self_reported_excel(excel_path):
    """Parse the self-reported leaderboard Excel into a DataFrame."""
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path))
    all_rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [
            cell.value
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, values_only=True
        ):
            if row[0] is None:
                continue
            row_dict = {"modality": sheet_name}
            for i, h in enumerate(headers):
                if h is not None and i < len(row):
                    row_dict[h] = row[i]
            all_rows.append(row_dict)

    return pd.DataFrame(all_rows)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def ensure_raw_data():
    """Download Excel files and task list."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    print("Step 1: Downloading metadata files...")
    download_file(
        EXCEL_VERIFIED_URL, RAW_DIR / "osworld_verified_results.xlsx"
    )
    download_file(
        EXCEL_SELF_REPORTED_URL, RAW_DIR / "self_reported_results.xlsx"
    )
    download_file(TASK_LIST_URL, RAW_DIR / "test_all.json")


def _extract_single_zip(args):
    """Worker function: extract results from one zip.

    Returns (model_key, combined_scores_dict, n_runs, error_msg).
    """
    zip_name, model_label, steps, cache_dir = args
    model_key = f"{model_label}__{steps}steps"
    safe_name = model_key.replace(" ", "_").replace("/", "_")
    cache_file = cache_dir / f"{safe_name}.json"

    # Check cache first
    if cache_file.exists():
        with open(cache_file, "r") as f:
            cached = json.load(f)
        return (model_key, cached, -1, None)  # -1 = from cache

    zip_url = HF_BASE + zip_name
    runs = extract_results_from_zip(zip_url)

    if not runs:
        return (model_key, {}, 0, f"No results in {zip_name}")

    # Average across runs if multiple
    if len(runs) == 1:
        combined = runs[0]
    else:
        all_keys = set()
        for run in runs:
            all_keys.update(run.keys())
        combined = {}
        for key in all_keys:
            scores = [r[key] for r in runs if key in r]
            combined[key] = sum(scores) / len(scores)

    # Cache
    with open(cache_file, "w") as f:
        json.dump(combined, f, indent=2)

    return (model_key, combined, len(runs), None)


def extract_all_results(max_workers=3):
    """Extract per-task results from all zip files via range requests.

    Uses parallel workers to speed up extraction.

    Returns:
        all_results: dict of {model_label_steps: {task_key: score}}
    """
    all_results = {}
    cache_dir = RAW_DIR / "extracted_results"
    cache_dir.mkdir(exist_ok=True)

    total = len(ZIP_MODEL_MAP)
    work_items = [
        (zn, ml, st, cache_dir)
        for zn, ml, st in ZIP_MODEL_MAP
    ]

    completed = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_zip = {
            executor.submit(_extract_single_zip, item): item
            for item in work_items
        }

        for future in as_completed(future_to_zip):
            completed += 1
            item = future_to_zip[future]
            zip_name = item[0]

            try:
                model_key, combined, n_runs, error = \
                    future.result()
            except Exception as exc:
                print(
                    f"  [{completed}/{total}] {zip_name}: "
                    f"EXCEPTION - {exc}"
                )
                continue

            elapsed = time.time() - start_time
            rate = completed / elapsed if elapsed > 0 else 0
            eta = (total - completed) / rate if rate > 0 else 0

            if error:
                print(
                    f"  [{completed}/{total}] {model_key}: "
                    f"WARNING - {error} "
                    f"[{elapsed:.0f}s, ETA {eta:.0f}s]"
                )
                continue

            if n_runs == -1:
                print(
                    f"  [{completed}/{total}] {model_key}: "
                    f"cache ({len(combined)} tasks) "
                    f"[{elapsed:.0f}s]"
                )
            else:
                print(
                    f"  [{completed}/{total}] {model_key}: "
                    f"{len(combined)} tasks from {n_runs} run(s) "
                    f"[{elapsed:.0f}s, ETA {eta:.0f}s]"
                )

            if model_key not in all_results:
                all_results[model_key] = combined
            else:
                for k, v in combined.items():
                    if k in all_results[model_key]:
                        existing = all_results[model_key][k]
                        all_results[model_key][k] = \
                            (existing + v) / 2
                    else:
                        all_results[model_key][k] = v

    return all_results


def build_response_matrix(all_results, task_list):
    """Build response matrix from extracted results.

    Returns:
        response_df: DataFrame with model keys as rows, task UUIDs as columns
        task_meta: DataFrame mapping task_id to domain
    """
    # Build flat task list: domain/uuid -> column index
    task_ids = []
    task_domains = []
    for domain in sorted(task_list.keys()):
        for uuid in sorted(task_list[domain]):
            task_key = f"{domain}/{uuid}"
            task_ids.append(task_key)
            task_domains.append(domain)

    # Build matrix
    rows = {}
    for model_key in sorted(all_results.keys()):
        scores = all_results[model_key]
        row = [scores.get(tid, np.nan) for tid in task_ids]
        rows[model_key] = row

    response_df = pd.DataFrame.from_dict(
        rows, orient="index", columns=task_ids
    )
    response_df.index.name = "model"

    task_meta = pd.DataFrame({
        "task_id": task_ids,
        "domain": task_domains,
        "uuid": [tid.split("/")[1] for tid in task_ids],
    })

    return response_df, task_meta


def build_binary_matrix(response_matrix):
    """Convert continuous scores to binary: 1 if score >= 0.99, else 0.
    Preserves NaN for unevaluated tasks."""
    binary = response_matrix.copy()
    mask = binary.notna()
    binary[mask] = (binary[mask] >= 0.99).astype(float)
    return binary


def build_model_summary(response_matrix, binary_matrix):
    """Build per-model summary statistics."""
    rows = []
    for model in response_matrix.index:
        scores = response_matrix.loc[model]
        binary = binary_matrix.loc[model]

        n_evaluated = int(scores.notna().sum())
        n_total = len(scores)
        avg_score = float(scores.mean()) if n_evaluated > 0 else 0.0
        perfect = int((binary == 1.0).sum())
        zero = int((scores == 0.0).sum())
        partial = n_evaluated - perfect - zero

        # Parse model name and steps from key
        parts = model.rsplit("__", 1)
        model_name = parts[0]
        step_str = parts[1] if len(parts) > 1 else ""
        steps = int(step_str.replace("steps", "")) if step_str else 0

        rows.append({
            "model_key": model,
            "model_name": model_name,
            "max_steps": steps,
            "tasks_evaluated": n_evaluated,
            "tasks_total": n_total,
            "coverage_pct": round(n_evaluated / n_total * 100, 1),
            "avg_score": round(avg_score, 4),
            "avg_score_pct": round(avg_score * 100, 2),
            "success_count": perfect,
            "success_rate_pct": (
                round(perfect / n_evaluated * 100, 2)
                if n_evaluated > 0 else 0.0
            ),
            "zero_score_tasks": zero,
            "partial_score_tasks": partial,
        })

    summary = pd.DataFrame(rows)
    summary = summary.sort_values(
        "avg_score", ascending=False
    ).reset_index(drop=True)
    return summary


def print_statistics(response_matrix, binary_matrix, model_summary,
                     task_meta):
    """Print comprehensive summary statistics."""
    n_models, n_tasks = response_matrix.shape
    total_cells = n_models * n_tasks
    n_evaluated = int(response_matrix.notna().sum().sum())
    n_missing = total_cells - n_evaluated
    fill_rate = n_evaluated / total_cells if total_cells > 0 else 0

    print("\n" + "=" * 70)
    print("  RESPONSE MATRIX STATISTICS")
    print("=" * 70)
    print(f"  Models (rows):         {n_models}")
    print(f"  Tasks (columns):       {n_tasks}")
    print(f"  Matrix dimensions:     {n_models} x {n_tasks}")
    print(f"  Total cells:           {total_cells:,}")
    print(
        f"  Evaluated cells:       {n_evaluated:,} "
        f"({fill_rate * 100:.1f}%)"
    )
    print(
        f"  Missing cells (NaN):   {n_missing:,} "
        f"({n_missing / total_cells * 100:.1f}%)"
    )

    # Score distribution
    vals = response_matrix.values[~np.isnan(response_matrix.values)]
    if len(vals) > 0:
        print(
            f"\n  Score statistics (evaluated cells):"
        )
        print(
            f"    Mean:    {vals.mean():.4f} "
            f"({vals.mean() * 100:.1f}%)"
        )
        print(f"    Median:  {np.median(vals):.4f}")
        print(f"    Std:     {vals.std():.4f}")
        n_zero = (vals == 0.0).sum()
        n_one = (vals >= 0.99).sum()
        n_partial = len(vals) - n_zero - n_one
        print(
            f"    Score=0: {n_zero:,} ({n_zero / len(vals) * 100:.1f}%)"
        )
        print(
            f"    Score=1: {n_one:,} ({n_one / len(vals) * 100:.1f}%)"
        )
        print(
            f"    Partial: {n_partial:,} "
            f"({n_partial / len(vals) * 100:.1f}%)"
        )

    # Per-model stats
    per_model_avg = response_matrix.mean(axis=1)
    if len(per_model_avg) > 0:
        best = per_model_avg.idxmax()
        worst = per_model_avg.idxmin()
        print(f"\n  Per-model average score:")
        print(
            f"    Best:   {per_model_avg.max() * 100:.1f}% ({best})"
        )
        print(
            f"    Worst:  {per_model_avg.min() * 100:.1f}% ({worst})"
        )
        print(
            f"    Median: {per_model_avg.median() * 100:.1f}%"
        )
        print(f"    Std:    {per_model_avg.std() * 100:.1f}%")

    # Per-task stats
    per_task_avg = response_matrix.mean(axis=0)
    if len(per_task_avg.dropna()) > 0:
        print(f"\n  Per-task solve rate:")
        print(f"    Min:    {per_task_avg.min() * 100:.1f}%")
        print(f"    Max:    {per_task_avg.max() * 100:.1f}%")
        print(f"    Median: {per_task_avg.median() * 100:.1f}%")
        print(f"    Std:    {per_task_avg.std() * 100:.1f}%")

    # Task difficulty (binary)
    per_task_pass = binary_matrix.mean(axis=0)
    unsolved = int((per_task_pass == 0).sum())
    easy = int((per_task_pass > 0.8).sum())
    hard = int(((per_task_pass > 0) & (per_task_pass <= 0.2)).sum())
    solved_all = int((per_task_pass == 1.0).sum())
    print(f"\n  Task difficulty (binary pass/fail):")
    print(f"    Solved by NO model:    {unsolved}")
    print(f"    Hard (<=20%):          {hard}")
    print(f"    Easy (>80%):           {easy}")
    print(f"    Solved by ALL models:  {solved_all}")

    # Domain breakdown
    print(f"\n  Domain breakdown:")
    for domain in sorted(task_meta["domain"].unique()):
        domain_tasks = task_meta[
            task_meta["domain"] == domain
        ]["task_id"].tolist()
        domain_cols = [
            c for c in response_matrix.columns if c in domain_tasks
        ]
        if domain_cols:
            domain_vals = response_matrix[domain_cols].values
            domain_flat = domain_vals[~np.isnan(domain_vals)]
            if len(domain_flat) > 0:
                avg = domain_flat.mean()
                perfect = (domain_flat >= 0.99).mean()
                print(
                    f"    {domain:22s}: {len(domain_cols):3d} tasks, "
                    f"avg={avg * 100:.1f}%, "
                    f"pass={perfect * 100:.1f}%"
                )

    # Top models
    print(f"\n  Top 15 models (by average score):")
    print(
        f"  {'Rank':<5} {'Model':<50} "
        f"{'Avg%':>6} {'Pass':>6} {'Tasks':>6}"
    )
    print(f"  {'-' * 5} {'-' * 50} {'-' * 6} {'-' * 6} {'-' * 6}")
    for i, (_, row) in enumerate(model_summary.head(15).iterrows()):
        print(
            f"  {i + 1:<5} {row['model_key']:<50} "
            f"{row['avg_score_pct']:>5.1f}% "
            f"{row['success_count']:>5d} "
            f"{row['tasks_evaluated']:>5d}"
        )


def main():
    print("OSWorld Response Matrix Builder")
    print("=" * 70)

    # Step 1: Download metadata
    ensure_raw_data()

    # Step 2: Parse leaderboard Excel files
    print("\nStep 2: Parsing leaderboard Excel files...")
    verified_path = RAW_DIR / "osworld_verified_results.xlsx"
    self_reported_path = RAW_DIR / "self_reported_results.xlsx"

    verified_df = parse_verified_excel(verified_path)
    self_reported_df = parse_self_reported_excel(self_reported_path)

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    verified_csv = PROCESSED_DIR / "leaderboard_verified.csv"
    verified_df.to_csv(verified_csv, index=False)
    print(f"  Saved verified leaderboard: {verified_csv}")
    print(
        f"  ({len(verified_df)} entries, "
        f"{verified_df['Model'].nunique()} unique models)"
    )

    self_reported_csv = PROCESSED_DIR / "leaderboard_self_reported.csv"
    self_reported_df.to_csv(self_reported_csv, index=False)
    print(f"  Saved self-reported leaderboard: {self_reported_csv}")
    print(
        f"  ({len(self_reported_df)} entries across "
        f"{self_reported_df['modality'].nunique()} modalities)"
    )

    # Step 3: Load task list
    print("\nStep 3: Loading task list...")
    task_list = load_task_list()
    total_tasks = sum(len(v) for v in task_list.values())
    print(f"  {total_tasks} tasks across {len(task_list)} domains:")
    for domain in sorted(task_list.keys()):
        print(f"    {domain}: {len(task_list[domain])} tasks")

    # Step 4: Extract per-task results from zip archives
    print(
        f"\nStep 4: Extracting per-task results from "
        f"{len(ZIP_MODEL_MAP)} zip archives..."
    )
    print(
        "  (Using HTTP range requests - no full downloads needed)"
    )
    all_results = extract_all_results()
    print(f"\n  Extracted results for {len(all_results)} model configs")

    # Step 5: Build response matrix
    print("\nStep 5: Building response matrices...")
    response_matrix, task_meta = build_response_matrix(
        all_results, task_list
    )
    binary_matrix = build_binary_matrix(response_matrix)
    print(f"  Matrix shape: {response_matrix.shape}")

    model_summary = build_model_summary(response_matrix, binary_matrix)

    # Step 6: Save outputs
    print("\nStep 6: Saving outputs...")

    matrix_path = PROCESSED_DIR / "response_matrix.csv"
    response_matrix.to_csv(matrix_path)
    print(f"  Saved: {matrix_path}")

    binary_path = PROCESSED_DIR / "response_matrix_binary.csv"
    binary_matrix.to_csv(binary_path)
    print(f"  Saved: {binary_path}")

    summary_path = PROCESSED_DIR / "model_summary.csv"
    model_summary.to_csv(summary_path, index=False)
    print(f"  Saved: {summary_path}")

    meta_path = PROCESSED_DIR / "task_metadata.csv"
    task_meta.to_csv(meta_path, index=False)
    print(f"  Saved: {meta_path}")

    # Step 7: Print statistics
    print_statistics(
        response_matrix, binary_matrix, model_summary, task_meta
    )

    # Final file listing
    print(f"\n{'=' * 70}")
    print("  OUTPUT FILES")
    print(f"{'=' * 70}")
    for f in sorted(PROCESSED_DIR.iterdir()):
        size_kb = f.stat().st_size / 1024
        print(f"  {f.name:45s}  {size_kb:.1f} KB")

    print("\nDone.")


if __name__ == "__main__":
    main()
